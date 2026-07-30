import streamlit as st
import pandas as pd
from collections import Counter
import sqlite3
import datetime
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# --- PAGE CONFIGURATION ---
st.set_page_config(page_title="Multi-Style Stock Variance Dashboard", layout="wide")
st.title("📊 Daily Multi-Style Stock Variance Dashboard")

# --- DATABASE SETUP ---
DB_NAME = "variance_history.db"

def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS daily_variance (
            date TEXT,
            style_number TEXT,
            location TEXT,
            color_size_code TEXT,
            color_size_name TEXT,
            color TEXT,
            size TEXT,
            system_qty INTEGER,
            physical_qty INTEGER,
            variance INTEGER
        )
    ''')
    conn.commit()
    conn.close()

init_db()

# --- HELPER FUNCTIONS ---
def extract_style_number(code_or_name):
    str_val = str(code_or_name).strip()
    if len(str_val) >= 8 and str_val[:8].isdigit():
        return str_val[:8]
    return "NOT EXIST IN ERP"

def parse_color_size(name):
    name = str(name)
    if name in ['nan', 'None', '']:
        return 'NOT EXIST IN ERP', 'NOT EXIST IN ERP'
    if '-' in name:
        parts = name.rsplit('-', 1)
        size = parts[-1].strip()
        color_part = parts[0].split('-')[-1].strip() if len(parts[0].split('-')) > 1 else 'NOT EXIST IN ERP'
        return color_part, size
    return 'NOT EXIST IN ERP', 'NOT EXIST IN ERP'

def process_scanned_data(file_or_text):
    if file_or_text is None:
        return []
    
    codes = []
    if hasattr(file_or_text, 'name'):
        filename = file_or_text.name.lower()
        if filename.endswith(('.xlsx', '.xls')):
            df_temp = pd.read_excel(file_or_text, header=None)
            codes = df_temp.iloc[:, 0].dropna().astype(str).str.strip().str.upper().tolist()
        elif filename.endswith('.csv'):
            df_temp = pd.read_csv(file_or_text, header=None)
            codes = df_temp.iloc[:, 0].dropna().astype(str).str.strip().str.upper().tolist()
        elif filename.endswith('.txt'):
            stringio = io.StringIO(file_or_text.getvalue().decode("utf-8"))
            codes = [line.strip().upper() for line in stringio.readlines() if line.strip()]
    elif isinstance(file_or_text, str):
        codes = [line.strip().upper() for line in file_or_text.split('\n') if line.strip()]
        
    return codes

def generate_formatted_excel(df, date_str, outlet_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Balance"
    ws.views.sheetView[0].showGridLines = True

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    bold_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    white_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # 1. Title Banner
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = f"STOCK BALANCE AS AT {date_str} {outlet_name.upper()}"
    title_cell.font = bold_font
    title_cell.fill = yellow_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 2. Table Headers
    headers = ["Style No", "COLOR", "System QTY", "Physical QTY", "VARIANCE"]
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h_text)
        cell.font = bold_font
        cell.fill = yellow_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 3. Data Rows
    current_row = 3
    for _, row in df.iterrows():
        style_no = str(row['Style No'])
        color = str(row['Color'])
        sys_qty = int(row['System Qty'])
        phys_qty = int(row['Physical Qty'])
        variance = int(row['Variance'])

        c1 = ws.cell(row=current_row, column=1, value=style_no)
        c2 = ws.cell(row=current_row, column=2, value=color)
        c3 = ws.cell(row=current_row, column=3, value=sys_qty)
        c4 = ws.cell(row=current_row, column=4, value=phys_qty)
        c5 = ws.cell(row=current_row, column=5, value=variance)

        c1.alignment = Alignment(horizontal="center", vertical="center")
        c2.alignment = Alignment(horizontal="left", vertical="center")
        c3.alignment = Alignment(horizontal="right", vertical="center")
        c4.alignment = Alignment(horizontal="right", vertical="center")
        c5.alignment = Alignment(horizontal="right", vertical="center")

        for c in [c1, c2, c3, c4, c5]:
            c.font = normal_font
            c.border = thin_border

        if variance < 0:
            for c in [c2, c3, c4, c5]:
                c.fill = red_fill
                c.font = white_font

        current_row += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# --- SIDEBAR CONTROLS ---
st.sidebar.title("🛠️ Daily Inputs")

st.sidebar.subheader("📅 Date Selection")
selected_date = st.sidebar.date_input("Report Date", datetime.date.today())

st.sidebar.subheader("📂 1. Upload Style ERP Files (e.g. 5 files)")
erp_files = st.sidebar.file_uploader(
    "Upload today's ERP Style Files", 
    type=["xlsx", "xls"], 
    accept_multiple_files=True,
    key="erp_multi"
)

st.sidebar.subheader("📱 2. Outlet Daily Scanned Files")
outlets = ["CCC", "COLOMBO 03", "NUGEGODA", "ONLINE", "WATTALA"]
outlet_scans = {}

for outlet in outlets:
    with st.sidebar.expander(f"📍 {outlet} (1 Combined File for All Styles)"):
        tab1, tab2 = st.tabs(["📁 Upload File", "📋 Paste Codes"])
        with tab1:
            u_file = st.file_uploader(f"Upload scanned file for {outlet}", type=["xlsx", "xls", "csv", "txt"], key=f"f_{outlet}")
        with tab2:
            p_text = st.text_area(f"Paste all barcodes for {outlet}:", height=80, key=f"t_{outlet}")
        
        if u_file:
            c = process_scanned_data(u_file)
            outlet_scans[outlet] = c
            st.caption(f"✓ {len(c)} total barcodes loaded from file")
        elif p_text:
            c = process_scanned_data(p_text)
            outlet_scans[outlet] = c
            st.caption(f"✓ {len(c)} total barcodes loaded from text")

st.sidebar.markdown("---")

st.sidebar.subheader("⚙️ 3. Process Report")
run_calc = st.sidebar.button("💾 Calculate & Save Daily Report", type="primary", use_container_width=True)

# Admin Database Cleanup Section
st.sidebar.markdown("---")
st.sidebar.subheader("🗑️ Admin: Reset Database")
if st.sidebar.button("⚠️ Clear All Saved Reports"):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("DELETE FROM daily_variance")
    conn.commit()
    conn.close()
    st.sidebar.success("Database cleared!")
    st.rerun()


# --- PROCESSING ACTION ---
if run_calc:
    if not erp_files:
        st.error("⚠️ Please upload at least one Raysoft ERP file in Step 1.")
    else:
        all_erp_dfs = []
        for f in erp_files:
            df_temp = pd.read_excel(f)
            df_temp.columns = df_temp.columns.str.strip()
            if 'Unnamed: 0' in df_temp.columns:
                df_temp = df_temp.drop(columns=['Unnamed: 0'])
            all_erp_dfs.append(df_temp)
            
        master_erp = pd.concat(all_erp_dfs, ignore_index=True)
        master_erp.rename(columns={'Qty': 'System Qty'}, inplace=True)
        
        master_erp['Color Size Code'] = master_erp['Color Size Code'].astype(str).str.strip().str.upper()
        
        master_erp['Style Number'] = master_erp['Color Size Code'].apply(extract_style_number)
        color_size_parsed = master_erp['Color Size Name'].apply(parse_color_size)
        master_erp['Color'] = [x[0] for x in color_size_parsed]
        master_erp['Size'] = [x[1] for x in color_size_parsed]

        # Aggregate Scans
        scanned_records = []
        for loc, codes in outlet_scans.items():
            if codes:
                counts = Counter(codes)
                for code, count in counts.items():
                    scanned_records.append({
                        'Location': loc,
                        'Color Size Code': code,
                        'Physical Qty': count
                    })
        
        scanned_df = pd.DataFrame(scanned_records)

        # Merge System vs Scanned Qty
        if not scanned_df.empty:
            merged = pd.merge(master_erp, scanned_df, on=['Location', 'Color Size Code'], how='outer')
        else:
            merged = master_erp.copy()
            merged['Physical Qty'] = 0

        merged['System Qty'] = merged['System Qty'].fillna(0)
        merged['Physical Qty'] = merged['Physical Qty'].fillna(0)
        merged['Variance'] = merged['Physical Qty'] - merged['System Qty']

        # CLEARLY MARK NON-EXISTENT SCANNED BARCODES
        erp_codes_set = set(master_erp['Color Size Code'].unique())

        merged['Style Number'] = merged.apply(
            lambda r: 'NOT EXIST IN ERP' if r['Color Size Code'] not in erp_codes_set else r['Style Number'], axis=1
        )
        merged['Color'] = merged.apply(
            lambda r: 'NOT EXIST IN ERP' if r['Color Size Code'] not in erp_codes_set else r['Color'], axis=1
        )
        merged['Size'] = merged.apply(
            lambda r: 'NOT EXIST IN ERP' if r['Color Size Code'] not in erp_codes_set else r['Size'], axis=1
        )
        merged['Color Size Name'] = merged.apply(
            lambda r: '⚠️ ITEM DOES NOT EXIST IN ERP MASTER' if r['Color Size Code'] not in erp_codes_set else r['Color Size Name'], axis=1
        )

        # Save to SQLite Database
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("DELETE FROM daily_variance WHERE date = ?", (str(selected_date),))
        
        insert_rows = []
        for _, r in merged.iterrows():
            insert_rows.append((
                str(selected_date),
                str(r.get('Style Number', extract_style_number(r['Color Size Code']))),
                str(r.get('Location', '')),
                str(r['Color Size Code']),
                str(r.get('Color Size Name', '')),
                str(r.get('Color', '')),
                str(r.get('Size', '')),
                int(r['System Qty']),
                int(r['Physical Qty']),
                int(r['Variance'])
            ))
        
        c.executemany('''
            INSERT INTO daily_variance VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', insert_rows)
        conn.commit()
        conn.close()
        
        st.success(f"🎉 Report for {selected_date} successfully processed and saved!")


# --- DASHBOARD AREA ---
st.header("🏢 Outlet Summary & Variance Analysis")

conn = sqlite3.connect(DB_NAME)
available_dates = pd.read_sql_query("SELECT DISTINCT date FROM daily_variance ORDER BY date DESC", conn)

if not available_dates.empty:
    col_d, col_o = st.columns(2)
    view_date = col_d.selectbox("Select Date to View", available_dates['date'].tolist())
    view_outlet = col_o.selectbox("Select Outlet Name", outlets)

    query = f"""
        SELECT style_number as 'Style No', 
               color as 'Color',
               size as 'Size',
               color_size_code as 'Color Size Code', 
               color_size_name as 'Description', 
               system_qty as 'System Qty', 
               physical_qty as 'Physical Qty', 
               variance as 'Variance'
        FROM daily_variance 
        WHERE date = '{view_date}' AND location = '{view_outlet}'
    """
    outlet_df = pd.read_sql_query(query, conn)
    conn.close()

    if not outlet_df.empty:
        # High-level metrics
        total_sys = outlet_df['System Qty'].sum()
        total_phys = outlet_df['Physical Qty'].sum()
        total_var = outlet_df['Variance'].sum()

        m1, m2, m3 = st.columns(3)
        m1.metric("Total System Qty", total_sys)
        m2.metric("Total Physical Qty", total_phys)
        m3.metric("Overall Net Variance", total_var, delta_color="inverse")

        # ALERT FOR NON-EXISTENT BARCODES
        unmapped_items = outlet_df[outlet_df['Style No'] == 'NOT EXIST IN ERP']
        if not unmapped_items.empty:
            st.warning(
                f"⚠️ **DETECTION ALERT:** Found **{len(unmapped_items)} barcode(s)** scanned at **{view_outlet}** "
                f"that DO NOT exist in any uploaded ERP Style file!"
            )
            with st.expander("🚨 Click here to inspect non-existent barcodes"):
                st.dataframe(
                    unmapped_items[['Color Size Code', 'Physical Qty', 'Description']],
                    use_container_width=True
                )

        st.markdown("---")

        # --- SUMMARY TABLE ---
        st.subheader(f"📌 Overall Summary for {view_outlet} ({view_date})")

        color_summary = outlet_df.groupby(['Style No', 'Color']).agg(
            System_Qty=('System Qty', 'sum'),
            Physical_Qty=('Physical Qty', 'sum'),
            Variance=('Variance', 'sum')
        ).reset_index()

        color_summary.rename(columns={
            'System_Qty': 'System Qty',
            'Physical_Qty': 'Physical Qty'
        }, inplace=True)

        st.dataframe(
            color_summary.style.map(
                lambda v: 'background-color: #ffcccc; color: #900c3f; font-weight: bold;' if v < 0 else ('background-color: #fff3cd; color: #856404; font-weight: bold;' if v > 0 else ''),
                subset=['Variance']
            ),
            use_container_width=True
        )

        formatted_excel_bytes = generate_formatted_excel(color_summary, str(view_date), view_outlet)
        st.download_button(
            label=f"📊 Download Formatted Excel Report ({view_outlet})",
            data=formatted_excel_bytes,
            file_name=f"STOCK_BALANCE_{view_outlet}_{view_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )

        st.markdown("---")

        # --- FILTERS & DETAILED BREAKDOWN ---
        st.subheader("🔍 Filter & Inspect Color / Size Wise Details")

        f_col1, f_col2, f_col3 = st.columns(3)
        
        styles_list = ["All Styles"] + sorted([s for s in outlet_df['Style No'].dropna().unique() if s])
        colors_list = ["All Colors"] + sorted([c for c in outlet_df['Color'].dropna().unique() if c])
        sizes_list = ["All Sizes"] + sorted([sz for sz in outlet_df['Size'].dropna().unique() if sz])

        sel_style = f_col1.selectbox("Filter by Style No", styles_list)
        sel_color = f_col2.selectbox("Filter by Color", colors_list)
        sel_size = f_col3.selectbox("Filter by Size", sizes_list)

        filtered_df = outlet_df.copy()
        if sel_style != "All Styles":
            filtered_df = filtered_df[filtered_df['Style No'] == sel_style]
        if sel_color != "All Colors":
            filtered_df = filtered_df[filtered_df['Color'] == sel_color]
        if sel_size != "All Sizes":
            filtered_df = filtered_df[filtered_df['Size'] == sel_size]

        st.dataframe(
            filtered_df.style.map(
                lambda v: 'background-color: #ffcccc; color: #900c3f; font-weight: bold;' if v < 0 else ('background-color: #fff3cd; color: #856404; font-weight: bold;' if v > 0 else ''),
                subset=['Variance']
            ),
            use_container_width=True
        )

        csv_out = filtered_df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label=f"📥 Download Detailed CSV Data for {view_outlet}",
            data=csv_out,
            file_name=f"{view_outlet}_detailed_{view_date}.csv",
            mime="text/csv"
        )
    else:
        st.info(f"No records found for {view_outlet} on {view_date}.")
else:
    conn.close()
    st.info("👈 Please upload ERP style files and outlet scan files in the left sidebar, then click 'Calculate & Save Daily Report'.")
